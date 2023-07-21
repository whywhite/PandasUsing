from copy import copy
from openpyxl.drawing.image import Image
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Fill, Alignment, Protection, PatternFill
import pandas as pd


def merge_excel(source_folder, target_file):
    excel_files = []

    for root, dirs, files in os.walk(source_folder):
        if "DRP_Compare" in dirs:
            dirs.remove("DRP_Compare")
        for file in files:
            if file.endswith('xlsx'):
                file_path = os.path.join(root, file)
                excel_files.append(file_path)
    print(excel_files)

    # 创建目标工作簿
    with pd.ExcelWriter(target_file) as writer:
        for excel in excel_files:
            # 读取源工作簿的所有sheet
            xls = pd.ExcelFile(excel)

            for sheet_name in xls.sheet_names:
                # 读取源sheet数据
                df = pd.read_excel(xls, sheet_name=sheet_name)

                # 写入目标工作簿的对应sheet
                df.columns = df.columns.astype(str).str.replace('Unnamed:.*', '')
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    print('合并完成')


def insert_image(image_dir, excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active
    column_max = sheet.max_column
    # row = 1
    # 遍历图片文件夹，匹配格式为jpg，png的图片
    for filename in os.listdir(image_dir):
        if re.match(r'.*\.(jpg|png|jpeg|bmp|gif|tif|tiff)', filename, re.IGNORECASE):
            img_path = os.path.join(image_dir, filename)
            img = Image(img_path)
            cell_found = False
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=column_max).value == filename:
                    cell_found = True
                    break
            if cell_found:
                continue
            for i in range(1, sheet.max_row + 4):
                if all(sheet.cell(row=i, column=j).value is None or sheet.cell(row=i, column=j).value == "" for j in
                       range(1, column_max + 1)):
                    sheet.row_dimensions[i].height = 500
                    cell = sheet.cell(row=i, column=1)
                    cell.value = filename
                    if img.width > 500:
                        img.width = 500
                    if img.height > 500:
                        img.height = 500
                    sheet.add_image(img, f'B{i}')
                    break
            else:
                print(f"找不到图片 {filename}, 找不到空行.")
    wb.save(excel_file)
    print("图片插入完成.")


if __name__ == '__main__':
    source_folder = '/Users/daji123/Desktop/upload'
    target_file = '/Users/daji123/Desktop/1/Testsum.xlsx'
    merge_excel(source_folder, target_file)
